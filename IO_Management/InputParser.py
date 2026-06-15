# InputParser.py
# ---------------------------------------------------------------------------
# Ensure project root is in sys.path when running this file directly
import sys
from pathlib import Path

for parent in Path(__file__).resolve().parents:
    if (parent / "EngineCore").exists() and (parent / "Thermodynamics").exists():
        if str(parent) not in sys.path:
            sys.path.insert(0, str(parent))
        break

from parapy.core import Base, Input, Attribute
from IO_Management.input_schema import INPUT_SCHEMA
from EngineCore.material_database import MATERIAL_DB

# Module-level defaults extracted from schema
_DEFAULT_FLIGHT = {k: v["default"] for k, v in INPUT_SCHEMA["design_flight_conditions"].items()}
_DEFAULT_FEATURES = {k: v["default"] for k, v in INPUT_SCHEMA["engine_features"].items()}
_DEFAULT_GEOMETRY = {k: v["default"] for k, v in INPUT_SCHEMA["engine_geometry"].items()}
_DEFAULT_MATERIALS = {k: v["default"] for k, v in INPUT_SCHEMA["engine_materials"].items()}


class InputParser(Base):
    """
    KBE Input Parser for AeroEngine.
    Reads/writes Excel configuration files, performs engineering validation,
    and provides a Tkinter GUI for editing design parameters.
    """

    # Path to the .xlsx design file. Empty string = use defaults.
    filepath = Input("")

    @Attribute
    def raw_data(self):
        """Parsed content from xlsx, or empty dicts if filepath is empty/missing."""
        return self._load() if self.filepath else ({}, {}, {}, {})

    @Attribute
    def flight_conditions(self):
        return {**_DEFAULT_FLIGHT, **self.raw_data[0]}

    @Attribute
    def engine_features(self):
        return {**_DEFAULT_FEATURES, **self.raw_data[1]}

    @Attribute
    def engine_geometry(self):
        return {**_DEFAULT_GEOMETRY, **self.raw_data[2]}

    @Attribute
    def engine_materials(self):
        return {**_DEFAULT_MATERIALS, **self.raw_data[3]}

    @Attribute
    def validation_errors(self):
        return self.validate(
            self.flight_conditions,
            self.engine_features,
            self.engine_geometry,
            self.engine_materials,
        )

    # ------------------------------------------------------------------
    # Plain Methods (Imperative side-effects)
    # ------------------------------------------------------------------

    def _load(self):
        """Reads the xlsx at self.filepath, returns 4-tuple of dicts (flight, features, geometry, materials)."""
        import os
        from openpyxl import load_workbook

        flight = {}
        features = {}
        geometry = {}
        materials = {}

        if not self.filepath or not os.path.exists(self.filepath):
            return flight, features, geometry, materials

        try:
            wb = load_workbook(self.filepath, data_only=True)
            if "Design Inputs" not in wb.sheetnames:
                return flight, features, geometry, materials

            ws = wb["Design Inputs"]
            for row in ws.iter_rows(values_only=True):
                if not row or len(row) < 3:
                    continue
                key = row[0]
                if not isinstance(key, str):
                    continue
                key = key.strip()
                val = row[2]
                if val is None:
                    continue

                # Check which section this key belongs to
                if key in INPUT_SCHEMA["design_flight_conditions"]:
                    schema_item = INPUT_SCHEMA["design_flight_conditions"][key]
                    try:
                        if schema_item["type"] == "float":
                            flight[key] = float(val)
                        elif schema_item["type"] == "int":
                            flight[key] = int(val)
                        else:
                            flight[key] = str(val).strip()
                    except (ValueError, TypeError):
                        pass
                elif key in INPUT_SCHEMA["engine_features"]:
                    schema_item = INPUT_SCHEMA["engine_features"][key]
                    try:
                        if schema_item["type"] == "float":
                            features[key] = float(val)
                        elif schema_item["type"] == "int":
                            features[key] = int(val)
                        else:
                            features[key] = str(val).strip()
                    except (ValueError, TypeError):
                        pass
                elif key in INPUT_SCHEMA["engine_geometry"]:
                    schema_item = INPUT_SCHEMA["engine_geometry"][key]
                    try:
                        if schema_item["type"] == "float":
                            geometry[key] = float(val)
                        elif schema_item["type"] == "int":
                            geometry[key] = int(val)
                        else:
                            geometry[key] = str(val).strip()
                    except (ValueError, TypeError):
                        pass
                elif key in INPUT_SCHEMA["engine_materials"]:
                    materials[key] = str(val).strip()
            wb.close()
        except Exception as e:
            print(f"[InputParser] Error loading Excel file: {e}")

        return flight, features, geometry, materials

    def _save_data(self, filepath, flight, features, geometry, materials):
        """Helper to write stacked sections and material database to Excel."""
        import os
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()

        # Sheet 1: Design Inputs
        ws1 = wb.active
        ws1.title = "Design Inputs"

        bold_font = Font(bold=True)

        sections = [
            ("design_flight_conditions", "SECTION 0 — design_flight_conditions", flight),
            ("engine_features", "SECTION 1 — engine_features", features),
            ("engine_geometry", "SECTION 2 — engine_geometry", geometry),
            ("engine_materials", "SECTION 3 — engine_materials", materials)
        ]

        for section_key, section_header, section_data in sections:
            row_idx = ws1.max_row
            if row_idx > 1 or (row_idx == 1 and ws1.cell(row=1, column=1).value is not None):
                ws1.append([])  # Blank row before section

            ws1.append([section_header])
            ws1.cell(row=ws1.max_row, column=1).font = bold_font

            ws1.append(["Key", "Label", "Value", "Unit"])
            header_row_idx = ws1.max_row
            for col in range(1, 5):
                ws1.cell(row=header_row_idx, column=col).font = bold_font

            schema_section = INPUT_SCHEMA[section_key]
            for key in schema_section:
                label = schema_section[key]["label"]
                val = section_data.get(key, schema_section[key]["default"])
                unit = schema_section[key]["unit"]
                ws1.append([key, label, val, unit])

        # Sheet 2: Material Database
        ws2 = wb.create_sheet(title="Material Database")
        headers = [
            "Material Name",
            "Density [kg/m³]",
            "Yield Strength [MPa]",
            "Ultimate Tensile Strength [MPa]",
            "Elongation at Break [%]"
        ]
        ws2.append(headers)
        for col in range(1, len(headers) + 1):
            ws2.cell(row=1, column=col).font = bold_font

        for name, props in MATERIAL_DB.items():
            density = props.get("density", 0.0)
            ys = props.get("yield_stress", 0.0) / 1e6
            uts = props.get("ultimate_tensile_strength", 0.0) / 1e6
            elong = props.get("fracture_strain", 0.0) * 100.0
            ws2.append([name, density, ys, uts, elong])

        os.makedirs(os.path.dirname(os.path.abspath(filepath)), exist_ok=True)
        wb.save(filepath)
        wb.close()

    def save(self, filepath):
        """Writes Sheet 1 ("Design Inputs") with 4 sections stacked vertically,
        and Sheet 2 ("Material Database") from material_database.py. Overwrites if exists.
        """
        self._save_data(filepath, self.flight_conditions, self.engine_features, self.engine_geometry, self.engine_materials)

    def validate(self, flight, features, geometry, materials):
        """Validates all input categories against boundaries and relations.
        Returns list of error message strings.
        """
        errors = []

        # 1. Range checks
        sections = [
            ("design_flight_conditions", flight),
            ("engine_features", features),
            ("engine_geometry", geometry),
        ]

        for sec_name, data in sections:
            schema_sec = INPUT_SCHEMA[sec_name]
            for key, schema_item in schema_sec.items():
                if schema_item["type"] in ["float", "int"]:
                    val = data.get(key)
                    if val is None:
                        val = schema_item["default"]

                    try:
                        v_num = float(val)
                    except (ValueError, TypeError):
                        errors.append(f"{schema_item['label']}: value {val} is not a valid number")
                        continue

                    min_val = schema_item.get("min")
                    max_val = schema_item.get("max")

                    if min_val is not None and v_num < min_val:
                        errors.append(f"{schema_item['label']}: value {val} is out of range [{min_val}, {max_val}]")
                    elif max_val is not None and v_num > max_val:
                        errors.append(f"{schema_item['label']}: value {val} is out of range [{min_val}, {max_val}]")

        # 2. Efficiencies and recoveries <= 1.0 explicitly
        for key, val in features.items():
            if key.endswith("_eta") or key in ["IPR", "CCPR", "mech_eta"]:
                try:
                    v_num = float(val)
                    if v_num > 1.0:
                        lbl = INPUT_SCHEMA["engine_features"][key]["label"]
                        errors.append(f"{lbl} must be less than or equal to 1.0")
                except (ValueError, TypeError):
                    pass

        # 3. Materials check
        for key, val in materials.items():
            if val not in MATERIAL_DB:
                lbl = INPUT_SCHEMA["engine_materials"][key]["label"]
                errors.append(f"{lbl}: '{val}' is not in the material database")

        # 4. Spool tip vs spool length
        spool_tip_length = geometry.get("spool_tip_length", _DEFAULT_GEOMETRY["spool_tip_length"])
        spool_length = geometry.get("spool_length", _DEFAULT_GEOMETRY["spool_length"])
        try:
            if float(spool_tip_length) >= float(spool_length):
                errors.append("spool_tip_length must be less than spool_length")
        except (ValueError, TypeError):
            pass

        return errors

    def launch_gui(self, filepath=None):
        """Opens a Tkinter window with four tabs for input configuration."""
        import tkinter as tk
        from tkinter import ttk, filedialog, messagebox
        import os

        # Prompt user to load a pre-existing file if filepath is not already provided
        if not filepath:
            temp_root = tk.Tk()
            temp_root.withdraw()
            use_existing = messagebox.askyesno(
                "Load Pre-existing File",
                "Would you like to load a pre-existing design configuration (.xlsx) file?",
                parent=temp_root
            )
            if use_existing:
                selected_file = filedialog.askopenfilename(
                    parent=temp_root,
                    title="Select Pre-existing Design File",
                    filetypes=[("Excel Files", "*.xlsx")]
                )
                if selected_file:
                    old_fp = self.filepath
                    self.filepath = selected_file
                    test_flight, test_features, test_geometry, test_materials = self._load()
                    self.filepath = old_fp
                    
                    if test_flight or test_features or test_geometry or test_materials:
                        filepath = selected_file
                    else:
                        messagebox.showwarning(
                            "Invalid File Format",
                            "The selected file is not a valid design configuration file. Opening with defaults instead.",
                            parent=temp_root
                        )
            temp_root.destroy()

        # Load existing data if filepath is provided and exists
        loaded_flight, loaded_features, loaded_geometry, loaded_materials = ({}, {}, {}, {})
        gui_filepath = filepath or self.filepath
        if gui_filepath and os.path.exists(gui_filepath):
            old_fp = self.filepath
            self.filepath = gui_filepath
            loaded_flight, loaded_features, loaded_geometry, loaded_materials = self._load()
            self.filepath = old_fp

        root = tk.Tk()
        root.title("AeroEngine Design Configurator")
        root.geometry("850x650")
        root.configure(bg="#20232a")

        # Dark theme Styling
        style = ttk.Style(root)
        style.theme_use('clam')

        style.configure(".", background="#20232a", foreground="#ffffff", font=("Segoe UI", 10))
        style.configure("TNotebook", background="#20232a", borderwidth=0)
        style.configure("TNotebook.Tab", background="#282c34", foreground="#abb2bf", font=("Segoe UI", 10, "bold"), padding=[12, 6])
        style.map("TNotebook.Tab", background=[("selected", "#20232a")], foreground=[("selected", "#61dafb")])
        style.configure("TFrame", background="#20232a")
        style.configure("TLabel", background="#20232a", foreground="#ffffff")
        style.configure("TEntry", fieldbackground="#353b45", foreground="#ffffff", borderwidth=1, relief="flat")
        style.configure("TCombobox", fieldbackground="#353b45", foreground="#ffffff", selectbackground="#353b45")

        style.configure("Primary.TButton", background="#4a90e2", foreground="#ffffff", font=("Segoe UI", 10, "bold"), borderwidth=0, padding=8)
        style.map("Primary.TButton", background=[("active", "#357abd")])
        style.configure("Secondary.TButton", background="#4f5b66", foreground="#ffffff", font=("Segoe UI", 10), borderwidth=0, padding=8)
        style.map("Secondary.TButton", background=[("active", "#65737e")])

        # Header Title
        title_label = ttk.Label(root, text="AeroEngine KBE Input Parser", font=("Segoe UI", 16, "bold"), foreground="#61dafb")
        title_label.pack(pady=15)

        notebook = ttk.Notebook(root)
        notebook.pack(fill="both", expand=True, padx=15, pady=5)

        # We will store variables for each category
        vars_flight = {}
        vars_features = {}
        vars_geometry = {}
        vars_materials = {}

        material_choices = list(MATERIAL_DB.keys())

        # Tab 0: Flight Conditions
        tab_flight = ttk.Frame(notebook)
        notebook.add(tab_flight, text="Flight Conditions")

        # Tab 1: Cycle & Performance
        tab_features = ttk.Frame(notebook)
        notebook.add(tab_features, text="Cycle & Performance")

        # Tab 2: Geometry
        tab_geometry = ttk.Frame(notebook)
        notebook.add(tab_geometry, text="Geometry")

        # Tab 3: Materials
        tab_materials = ttk.Frame(notebook)
        notebook.add(tab_materials, text="Materials")

        # Populate Tab 0: Flight Conditions (Single column)
        flight_frame = ttk.Frame(tab_flight)
        flight_frame.pack(padx=20, pady=20, fill="both", expand=True)
        for i, (key, item) in enumerate(INPUT_SCHEMA["design_flight_conditions"].items()):
            val = loaded_flight.get(key, item["default"])
            var = tk.StringVar(value=str(val))
            vars_flight[key] = var

            lbl = ttk.Label(flight_frame, text=item["label"] + ":")
            lbl.grid(row=i, column=0, sticky="w", padx=10, pady=8)

            ent = ttk.Entry(flight_frame, textvariable=var, width=15)
            ent.grid(row=i, column=1, padx=10, pady=8, sticky="w")

            unit_text = f"{item['unit']}  [{item['min']} to {item['max']}]"
            unit_lbl = ttk.Label(flight_frame, text=unit_text, foreground="#abb2bf")
            unit_lbl.grid(row=i, column=2, padx=10, pady=8, sticky="w")

        # Populate Tab 1: Cycle & Performance (Two columns)
        features_frame = ttk.Frame(tab_features)
        features_frame.pack(padx=20, pady=20, fill="both", expand=True)
        feat_keys = list(INPUT_SCHEMA["engine_features"].keys())
        half_feat = (len(feat_keys) + 1) // 2

        for idx, key in enumerate(feat_keys):
            item = INPUT_SCHEMA["engine_features"][key]
            val = loaded_features.get(key, item["default"])
            var = tk.StringVar(value=str(val))
            vars_features[key] = var

            col_offset = 0 if idx < half_feat else 4
            row_idx = idx if idx < half_feat else idx - half_feat

            lbl = ttk.Label(features_frame, text=item["label"] + ":")
            lbl.grid(row=row_idx, column=col_offset, sticky="w", padx=5, pady=4)

            ent = ttk.Entry(features_frame, textvariable=var, width=10)
            ent.grid(row=row_idx, column=col_offset+1, padx=5, pady=4, sticky="w")

            unit_text = f"{item['unit']} [{item['min']}-{item['max']}]"
            unit_lbl = ttk.Label(features_frame, text=unit_text, foreground="#abb2bf", font=("Segoe UI", 9))
            unit_lbl.grid(row=row_idx, column=col_offset+2, padx=5, pady=4, sticky="w")

            # Spacer column between the two halves
            if col_offset == 0:
                spacer = ttk.Label(features_frame, text="   ")
                spacer.grid(row=row_idx, column=3)

        # Populate Tab 2: Geometry (Two columns)
        geom_frame = ttk.Frame(tab_geometry)
        geom_frame.pack(padx=20, pady=20, fill="both", expand=True)
        geom_keys = list(INPUT_SCHEMA["engine_geometry"].keys())
        half_geom = (len(geom_keys) + 1) // 2

        for idx, key in enumerate(geom_keys):
            item = INPUT_SCHEMA["engine_geometry"][key]
            val = loaded_geometry.get(key, item["default"])
            var = tk.StringVar(value=str(val))
            vars_geometry[key] = var

            col_offset = 0 if idx < half_geom else 4
            row_idx = idx if idx < half_geom else idx - half_geom

            lbl = ttk.Label(geom_frame, text=item["label"] + ":")
            lbl.grid(row=row_idx, column=col_offset, sticky="w", padx=8, pady=6)

            ent = ttk.Entry(geom_frame, textvariable=var, width=12)
            ent.grid(row=row_idx, column=col_offset+1, padx=8, pady=6, sticky="w")

            unit_text = f"{item['unit']} [{item['min']}-{item['max']}]"
            unit_lbl = ttk.Label(geom_frame, text=unit_text, foreground="#abb2bf")
            unit_lbl.grid(row=row_idx, column=col_offset+2, padx=8, pady=6, sticky="w")

            if col_offset == 0:
                spacer = ttk.Label(geom_frame, text="     ")
                spacer.grid(row=row_idx, column=3)

        # Populate Tab 3: Materials (Single column with Comboboxes)
        materials_frame = ttk.Frame(tab_materials)
        materials_frame.pack(padx=20, pady=20, fill="both", expand=True)
        for i, (key, item) in enumerate(INPUT_SCHEMA["engine_materials"].items()):
            val = loaded_materials.get(key, item["default"])
            var = tk.StringVar(value=str(val))
            vars_materials[key] = var

            lbl = ttk.Label(materials_frame, text=item["label"] + ":")
            lbl.grid(row=i, column=0, sticky="w", padx=10, pady=10)

            cb = ttk.Combobox(materials_frame, textvariable=var, values=material_choices, width=25, state="readonly")
            cb.grid(row=i, column=1, padx=10, pady=10, sticky="w")

            unit_lbl = ttk.Label(materials_frame, text=item["unit"], foreground="#abb2bf")
            unit_lbl.grid(row=i, column=2, padx=10, pady=10, sticky="w")

        # Error text area at the bottom
        bottom_frame = ttk.Frame(root)
        bottom_frame.pack(fill="x", side="bottom", padx=15, pady=15)

        error_title = ttk.Label(bottom_frame, text="Validation Errors:", font=("Segoe UI", 10, "bold"), foreground="#ff5555")
        error_title.pack(anchor="w")

        error_text = tk.Text(bottom_frame, height=5, fg="#ff5555", bg="#1e1e24", insertbackground="white", font=("Courier New", 9))
        error_text.pack(fill="x", pady=5)
        error_text.config(state="disabled")

        # Action Buttons frame
        buttons_frame = ttk.Frame(bottom_frame)
        buttons_frame.pack(anchor="e", pady=5)

        gui_result = [None]  # to store the result on success

        def on_save():
            # Clear previous errors
            error_text.config(state="normal")
            error_text.delete("1.0", tk.END)
            error_text.config(state="disabled")

            flight = {}
            features = {}
            geometry = {}
            materials = {}

            # Read variables and convert types
            conversion_errors = []

            # Flight Conditions
            for k, var in vars_flight.items():
                schema_item = INPUT_SCHEMA["design_flight_conditions"][k]
                raw_val = var.get().strip()
                try:
                    if schema_item["type"] == "float":
                        flight[k] = float(raw_val)
                    elif schema_item["type"] == "int":
                        flight[k] = int(raw_val)
                except ValueError:
                    conversion_errors.append(f"{schema_item['label']}: '{raw_val}' is not a valid number")
                    flight[k] = raw_val

            # Engine Features
            for k, var in vars_features.items():
                schema_item = INPUT_SCHEMA["engine_features"][k]
                raw_val = var.get().strip()
                try:
                    if schema_item["type"] == "float":
                        features[k] = float(raw_val)
                    elif schema_item["type"] == "int":
                        features[k] = int(raw_val)
                except ValueError:
                    conversion_errors.append(f"{schema_item['label']}: '{raw_val}' is not a valid number")
                    features[k] = raw_val

            # Engine Geometry
            for k, var in vars_geometry.items():
                schema_item = INPUT_SCHEMA["engine_geometry"][k]
                raw_val = var.get().strip()
                try:
                    if schema_item["type"] == "float":
                        geometry[k] = float(raw_val)
                    elif schema_item["type"] == "int":
                        geometry[k] = int(raw_val)
                except ValueError:
                    conversion_errors.append(f"{schema_item['label']}: '{raw_val}' is not a valid number")
                    geometry[k] = raw_val

            # Engine Materials
            for k, var in vars_materials.items():
                materials[k] = var.get().strip()

            # Run validator
            val_errors = self.validate(flight, features, geometry, materials)
            all_errors = conversion_errors + val_errors

            if all_errors:
                error_text.config(state="normal")
                for err in all_errors:
                    error_text.insert(tk.END, f"• {err}\n")
                error_text.config(state="disabled")
                return

            # If valid, ask for filename if gui_filepath is not set
            nonlocal gui_filepath
            if not gui_filepath:
                selected_path = filedialog.asksaveasfilename(
                    parent=root,
                    title="Save Design Configuration",
                    defaultextension=".xlsx",
                    filetypes=[("Excel Files", "*.xlsx")],
                    initialfile="engine_design.xlsx"
                )
                if not selected_path:
                    return
                gui_filepath = selected_path

            # Save the newly entered values to the Excel file
            self._save_data(gui_filepath, flight, features, geometry, materials)

            # Store the result to return
            gui_result[0] = (gui_filepath, flight, features, geometry, materials)
            root.destroy()

        def on_cancel():
            root.destroy()

        save_btn = ttk.Button(buttons_frame, text="Save & Close", command=on_save, style="Primary.TButton")
        save_btn.pack(side="left", padx=5)

        cancel_btn = ttk.Button(buttons_frame, text="Cancel", command=on_cancel, style="Secondary.TButton")
        cancel_btn.pack(side="left", padx=5)

        root.protocol("WM_DELETE_WINDOW", on_cancel)

        root.mainloop()

        return gui_result[0]


if __name__ == "__main__":
    parser = InputParser()
    print("Testing launch_gui with defaults...")
    # This will open the GUI, prefilled with defaults since filepath is empty.
    # When user clicks Save & Close, it will ask where to save.
    result = parser.launch_gui()
    print(f"GUI returned: {result}")

    if result:
        filepath = result[0]
        # Instantiate a new parser with the written path and verify values
        print(f"Loading written file: {filepath}")
        new_parser = InputParser(filepath=filepath)
        print(f"Mach condition: {new_parser.flight_conditions['Mach']} (Expected from GUI)")

        # Test validation on invalid inputs
        print("Testing validation on invalid inputs...")
        errors1 = new_parser.validate(
            flight={**new_parser.flight_conditions, "Mach": 1.5},  # out of range [0.0, 1.0]
            features={**new_parser.engine_features, "C_eta": 1.5},  # out of range [0.80, 1.0]
            geometry={**new_parser.engine_geometry, "spool_tip_length": 2.0, "spool_length": 1.5},  # tip >= length
            materials={**new_parser.engine_materials, "casing": "Unobtainium"}  # not in DB
        )
        print("Validation errors:")
        for err in errors1:
            print(f"  {err}")
