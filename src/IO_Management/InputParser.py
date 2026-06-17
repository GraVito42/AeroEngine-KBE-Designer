# InputParser.py
# ---------------------------------------------------------------------------
# Ensure project root is in sys.path when running this file directly
import sys
from pathlib import Path

for parent in Path(__file__).resolve().parents:
    if (parent / "EngineCore").exists() and (parent / "Thermodynamics").exists():
        if str(parent) not in sys.path:
            sys.path.insert(0, str(parent))
        break

from parapy.core import Base, Input, Attribute
from IO_Management.input_schema import INPUT_SCHEMA
from EngineCore.material_database import MATERIAL_DB

# Module-level defaults extracted from schema
_DEFAULT_FLIGHT = {k: v["default"] for k, v in INPUT_SCHEMA["design_flight_conditions"].items()}
_DEFAULT_FEATURES = {k: v["default"] for k, v in INPUT_SCHEMA["engine_features"].items()}
_DEFAULT_GEOMETRY = {k: v["default"] for k, v in INPUT_SCHEMA["engine_geometry"].items()}
_DEFAULT_MATERIALS = {k: v["default"] for k, v in INPUT_SCHEMA["engine_materials"].items()}

SESSION_FILENAME = "session_input.xlsx"


class InputParser(Base):
    """
    KBE Input Parser for AeroEngine.
    Reads/writes Excel configuration files, performs engineering validation,
    and provides a Tkinter GUI for editing design parameters.
    """

    # Path to the .xlsx design file. Empty string = use defaults.
    filepath = Input("")
    _on_save_callback = Input(None)

    # Plain Python attribute to store the selected working directory
    work_dir = ""

    @Attribute
    def raw_data(self):
        """Parsed content from xlsx, or empty dicts if filepath is empty/missing."""
        return self._load() if self.filepath else ({}, {}, {}, {})

    @Attribute
    def flight_conditions(self):
        return {**_DEFAULT_FLIGHT, **self.raw_data[0]}

    @Attribute
    def engine_features(self):
        return {**_DEFAULT_FEATURES, **self.raw_data[1]}

    @Attribute
    def engine_geometry(self):
        return {**_DEFAULT_GEOMETRY, **self.raw_data[2]}

    @Attribute
    def engine_materials(self):
        return {**_DEFAULT_MATERIALS, **self.raw_data[3]}

    @Attribute
    def compressor_n_stages(self):
        """Estimated compressor stage count from PR ceiling."""
        import math
        CPR    = self.engine_features["CPR"]
        PR_max = self.engine_features["stage_PR_max"]
        return max(1, math.ceil(math.log(CPR) / math.log(PR_max)))

    @Attribute
    def turbine_n_stages(self):
        """Estimated turbine stage count from shaft loading balance."""
        import math
        # ISA conditions at cruise altitude
        T0, _ = self._isa_conditions(
            self.flight_conditions["altitude"],
            self.flight_conditions.get("ISA_deviation", 0.0),
        )
        gamma_a = 1.4
        Tt2 = T0 * (1 + 0.5 * (gamma_a - 1) * self.flight_conditions["Mach"] ** 2)
        Tt3 = Tt2 * (1 + (1 / self.engine_features["C_eta"]) * (
            self.engine_features["CPR"] ** ((gamma_a - 1) / gamma_a) - 1))
        delta_h_c = 1000.0 * (Tt3 - Tt2)   # cp_air = 1000 J/kg/K

        gamma_g = self.engine_features["gamma_g"]
        cp_g    = self.engine_features["cpg"]
        Tt4     = self.engine_features["TIT"]
        # Power balance: mech_eta * delta_h_t = delta_h_c  (per unit mass flow)
        Tt5 = Tt4 - delta_h_c / (cp_g * self.engine_features["mech_eta"])
        delta_h_t = cp_g * (Tt4 - Tt5)

        n_c   = self.compressor_n_stages
        psi_c = self.engine_features["C_work_coeff"]
        psi_t = self.engine_features["T_work_coeff"]
        return max(1, math.ceil(delta_h_t * n_c * psi_c / (psi_t * delta_h_c)))

    @Attribute
    def validation_errors(self):
        return self.validate(
            self.flight_conditions,
            self.engine_features,
            self.engine_geometry,
            self.engine_materials,
        )

    # ------------------------------------------------------------------
    # Plain Methods (Imperative side-effects)
    # ------------------------------------------------------------------

    def configure_inputs(self):
        """Action/method to launch the GUI configuration."""
        return self.launch_gui()

    def validate_on_load(self):
        """Called by AeroEngine if filepath was set externally (no UI)."""
        errors = self.validate(
            self.flight_conditions,
            self.engine_features,
            self.engine_geometry,
            self.engine_materials,
        )
        level2 = [e for e in errors if e.startswith("ERROR")]
        for w in [e for e in errors if e.startswith("WARNING")]:
            print(f"[InputParser] {w}")
        if level2:
            raise RuntimeError(
                "[InputParser] Input file failed cross-parameter validation:\n" +
                "\n".join(f"  {e}" for e in level2)
            )

    def _load(self, path=None):
        """Reads the xlsx at target path, returns 4-tuple of dicts (flight, features, geometry, materials)."""
        import os
        from openpyxl import load_workbook

        target = path or self.filepath
        flight = {}
        features = {}
        geometry = {}
        materials = {}

        if not target or not os.path.exists(target):
            return flight, features, geometry, materials

        try:
            wb = load_workbook(target, data_only=True)
            if "Design Inputs" not in wb.sheetnames:
                return flight, features, geometry, materials

            ws = wb["Design Inputs"]
            for row in ws.iter_rows(values_only=True):
                if not row or len(row) < 3:
                    continue
                key = row[0]
                if not isinstance(key, str):
                    continue
                key = key.strip()
                val = row[2]
                if val is None:
                    continue

                # Check which section this key belongs to
                if key in INPUT_SCHEMA["design_flight_conditions"]:
                    schema_item = INPUT_SCHEMA["design_flight_conditions"][key]
                    try:
                        if schema_item["type"] == "float":
                            flight[key] = float(val)
                        elif schema_item["type"] == "int":
                            flight[key] = int(val)
                        else:
                            flight[key] = str(val).strip()
                    except (ValueError, TypeError):
                        pass
                elif key in INPUT_SCHEMA["engine_features"]:
                    schema_item = INPUT_SCHEMA["engine_features"][key]
                    try:
                        if schema_item["type"] == "float":
                            features[key] = float(val)
                        elif schema_item["type"] == "int":
                            features[key] = int(val)
                        else:
                            features[key] = str(val).strip()
                    except (ValueError, TypeError):
                        pass
                elif key in INPUT_SCHEMA["engine_geometry"]:
                    schema_item = INPUT_SCHEMA["engine_geometry"][key]
                    try:
                        if schema_item["type"] == "float":
                            geometry[key] = float(val)
                        elif schema_item["type"] == "int":
                            geometry[key] = int(val)
                        else:
                            geometry[key] = str(val).strip()
                    except (ValueError, TypeError):
                        pass
                elif key in INPUT_SCHEMA["engine_materials"]:
                    materials[key] = str(val).strip()
            wb.close()
        except Exception as e:
            print(f"[InputParser] Error loading Excel file: {e}")

        return flight, features, geometry, materials

    def _save_data(self, filepath, flight, features, geometry, materials):
        """Helper to write stacked sections and material database to Excel."""
        import os
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()

        # Sheet 1: Design Inputs
        ws1 = wb.active
        ws1.title = "Design Inputs"

        bold_font = Font(bold=True)

        sections = [
            ("design_flight_conditions", "SECTION 0 — design_flight_conditions", flight),
            ("engine_features", "SECTION 1 — engine_features", features),
            ("engine_geometry", "SECTION 2 — engine_geometry", geometry),
            ("engine_materials", "SECTION 3 — engine_materials", materials)
        ]

        for section_key, section_header, section_data in sections:
            row_idx = ws1.max_row
            if row_idx > 1 or (row_idx == 1 and ws1.cell(row=1, column=1).value is not None):
                ws1.append([])  # Blank row before section

            ws1.append([section_header])
            ws1.cell(row=ws1.max_row, column=1).font = bold_font

            ws1.append(["Key", "Label", "Value", "Unit"])
            header_row_idx = ws1.max_row
            for col in range(1, 5):
                ws1.cell(row=header_row_idx, column=col).font = bold_font

            schema_section = INPUT_SCHEMA[section_key]
            for key in schema_section:
                label = schema_section[key]["label"]
                val = section_data.get(key, schema_section[key]["default"])
                unit = schema_section[key]["unit"]
                ws1.append([key, label, val, unit])

        # Sheet 2: Material Database
        ws2 = wb.create_sheet(title="Material Database")
        headers = [
            "Material Name",
            "Density [kg/m³]",
            "Yield Strength [MPa]",
            "Ultimate Tensile Strength [MPa]",
            "Elongation at Break [%]"
        ]
        ws2.append(headers)
        for col in range(1, len(headers) + 1):
            ws2.cell(row=1, column=col).font = bold_font

        for name, props in MATERIAL_DB.items():
            density = props.get("density", 0.0)
            ys = props.get("yield_stress", 0.0) / 1e6
            uts = props.get("ultimate_tensile_strength", 0.0) / 1e6
            elong = props.get("fracture_strain", 0.0) * 100.0
            ws2.append([name, density, ys, uts, elong])

        os.makedirs(os.path.dirname(os.path.abspath(filepath)), exist_ok=True)
        wb.save(filepath)
        wb.close()

    def save(self, filepath):
        """Writes Sheet 1 ("Design Inputs") with 4 sections stacked vertically,
        and Sheet 2 ("Material Database") from material_database.py. Overwrites if exists.
        """
        self._save_data(filepath, self.flight_conditions, self.engine_features, self.engine_geometry, self.engine_materials)

    @staticmethod
    def _isa_conditions(altitude, isa_deviation=0.0):
        """Return (T0, p0) at the given altitude [m] using the ISA troposphere model.
        Includes optional ISA temperature deviation [K].
        Valid for altitudes 0–20 000 m (troposphere + lower stratosphere).
        """
        T_sl = 288.15   # Sea-level standard temperature  [K]
        p_sl = 101325.0  # Sea-level standard pressure    [Pa]
        L = 0.0065       # Temperature lapse rate          [K/m]
        g = 9.80665      # Gravitational acceleration      [m/s²]
        R = 287.05       # Specific gas constant (air)     [J/kg/K]

        if altitude <= 11000.0:
            T0 = T_sl - L * altitude + isa_deviation
            p0 = p_sl * ((T_sl - L * altitude) / T_sl) ** (g / (R * L))
        else:
            # Tropopause: isothermal layer above 11 km
            T_11 = T_sl - L * 11000.0
            p_11 = p_sl * (T_11 / T_sl) ** (g / (R * L))
            import math as _m
            T0 = T_11 + isa_deviation
            p0 = p_11 * _m.exp(-g * (altitude - 11000.0) / (R * T_11))

        return T0, p0

    def validate(self, flight, features, geometry, materials):
        """Validates all input categories against boundaries and relations.
        Returns list of error message strings.
        """
        import math

        errors = []

        # 1. Range checks
        sections = [
            ("design_flight_conditions", flight),
            ("engine_features", features),
            ("engine_geometry", geometry),
        ]

        for sec_name, data in sections:
            schema_sec = INPUT_SCHEMA[sec_name]
            for key, schema_item in schema_sec.items():
                if schema_item["type"] in ["float", "int"]:
                    val = data.get(key)
                    if val is None:
                        val = schema_item["default"]

                    try:
                        v_num = float(val)
                    except (ValueError, TypeError):
                        errors.append(f"{schema_item['label']}: value {val} is not a valid number")
                        continue

                    min_val = schema_item.get("min")
                    max_val = schema_item.get("max")

                    if min_val is not None and v_num < min_val:
                        errors.append(f"{schema_item['label']}: value {val} is out of range [{min_val}, {max_val}]")
                    elif max_val is not None and v_num > max_val:
                        errors.append(f"{schema_item['label']}: value {val} is out of range [{min_val}, {max_val}]")

        # 2. Efficiencies and recoveries <= 1.0 explicitly
        for key, val in features.items():
            if key.endswith("_eta") or key in ["IPR", "CCPR", "mech_eta"]:
                try:
                    v_num = float(val)
                    if v_num > 1.0:
                        lbl = INPUT_SCHEMA["engine_features"][key]["label"]
                        errors.append(f"{lbl} must be less than or equal to 1.0")
                except (ValueError, TypeError):
                    pass

        # 3. Materials check
        for key, val in materials.items():
            if val not in MATERIAL_DB:
                lbl = INPUT_SCHEMA["engine_materials"][key]["label"]
                errors.append(f"{lbl}: '{val}' is not in the material database")

        # ------------------------------------------------------------------ #
        # Cross-parameter feasibility checks
        # ------------------------------------------------------------------ #
        # Bail out early if any range/type errors were found — the cross-checks
        # depend on numeric values being valid.
        if errors:
            return errors

        # ---- CHECK 1: Wall thickness vs sheet thickness ---- #
        # sheet_thickness is now in input_schema; fallback 0.003 m for legacy files.
        sheet_thickness = float(geometry.get("sheet_thickness",
                                             INPUT_SCHEMA["engine_geometry"]["sheet_thickness"]["default"]))
        for key in ["casing_wall_thickness", "inlet_wall_thickness", "nozzle_wall_thickness"]:
            wall_val = geometry.get(key)
            if wall_val is None:
                wall_val = INPUT_SCHEMA["engine_geometry"][key]["default"]
            wall_val = float(wall_val)
            if wall_val <= 2 * sheet_thickness:
                errors.append(
                    f"ERROR: {key} = {wall_val:.4f} m must be > "
                    f"2 * sheet_thickness = {2 * sheet_thickness:.4f} m"
                )

        # ---- Common thermodynamic pre-computations (for Check 3) ---- #
        altitude = float(flight.get("altitude", _DEFAULT_FLIGHT["altitude"]))
        isa_dev  = float(flight.get("ISA_deviation", _DEFAULT_FLIGHT["ISA_deviation"]))
        M0       = float(flight.get("Mach", _DEFAULT_FLIGHT["Mach"]))

        T0_isa, p0_isa = self._isa_conditions(altitude, isa_dev)

        gamma_a  = 1.4
        cp_air   = 1000.0   # J/kg/K
        r_gas    = float(features.get("r_gas", _DEFAULT_FEATURES["r_gas"]))

        Tt0 = T0_isa * (1.0 + 0.5 * (gamma_a - 1.0) * M0 ** 2)
        Tt2 = Tt0  # Station 2 ≈ Station 0 total temperature

        CPR   = float(features.get("CPR", _DEFAULT_FEATURES["CPR"]))
        C_eta = float(features.get("C_eta", _DEFAULT_FEATURES["C_eta"]))
        T_eta = float(features.get("T_eta", _DEFAULT_FEATURES["T_eta"]))
        TIT   = float(features.get("TIT", _DEFAULT_FEATURES["TIT"]))
        mech_eta = float(features.get("mech_eta", _DEFAULT_FEATURES["mech_eta"]))

        # ---- CHECK 2: Stage count feasibility (via @Attribute) ---- #
        for name, n in [("Compressor", self.compressor_n_stages),
                        ("Turbine",    self.turbine_n_stages)]:
            if n >= 12:
                if name == "Compressor":
                    errors.append(
                        f"ERROR: Compressor n_stages={n} >= 12 — MEANGEN will crash (exit code 3).\n"
                        f"  n_stages = ceil(ln(CPR) / ln(stage_PR_max))\n"
                        f"  Current: CPR={self.engine_features['CPR']:.1f}, "
                        f"stage_PR_max={self.engine_features['stage_PR_max']:.3f}\n"
                        f"  → Reduce CPR (currently {self.engine_features['CPR']:.1f}) OR\n"
                        f"  → Increase stage_PR_max (currently {self.engine_features['stage_PR_max']:.3f})"
                    )
                else:
                    errors.append(
                        f"ERROR: Turbine n_stages={n} >= 12 — MEANGEN will crash (exit code 3).\n"
                        f"  n_stages = ceil(delta_h_t * n_c * C_work_coeff / (T_work_coeff * delta_h_c))\n"
                        f"  Current: TIT={self.engine_features['TIT']:.0f} K, "
                        f"T_work_coeff={self.engine_features['T_work_coeff']:.2f}, "
                        f"C_work_coeff={self.engine_features['C_work_coeff']:.2f}, "
                        f"CPR={self.engine_features['CPR']:.1f}\n"
                        f"  → Reduce TIT (currently {self.engine_features['TIT']:.0f} K) OR\n"
                        f"  → Increase T_work_coeff (currently {self.engine_features['T_work_coeff']:.2f}) OR\n"
                        f"  → Reduce C_work_coeff or CPR"
                    )
            elif n >= 8:
                print(
                    f"[InputParser] WARNING: {name} n_stages={n} is between 8 and 12. "
                    f"The model will open but CFD will likely crash above 10 stages.\n"
                    f"  Consider reducing "
                    f"{'CPR or increasing stage_PR_max' if name=='Compressor' else 'TIT or increasing T_work_coeff'} "
                    f"before running CFD."
                )

        # ---- CHECK 3: Mass flow vs d_max geometric compatibility ---- #
        try:
            from Thermodynamics.TurbojetSimplified import TurbojetSimplified

            model = TurbojetSimplified(
                target_thrust=float(features.get("Thrust_required",
                                                 _DEFAULT_FEATURES["Thrust_required"])),
                M0=M0, T0=T0_isa, p0=p0_isa,
                comp_pr=CPR, comp_eff=C_eta,
                Tt4=TIT, turb_eff=T_eta,
                inlet_pr=float(features.get("IPR", _DEFAULT_FEATURES["IPR"])),
                comb_pr=float(features.get("CCPR", _DEFAULT_FEATURES["CCPR"])),
                mech_eff=mech_eta,
            )
            mass_flow = model.mass_flow

            d_max = float(geometry.get("d_max", _DEFAULT_GEOMETRY["d_max"]))
            r_max = d_max / 2.0
            HTR = 0.4  # Hub-to-tip ratio default
            A_geom = math.pi * r_max ** 2 * (1.0 - HTR ** 2)

            # Flow area required from continuity at station 2 (Mach ≈ 0.5 axial)
            Mach2 = 0.5
            IPR = float(features.get("IPR", _DEFAULT_FEATURES["IPR"]))
            T2 = Tt2 / (1.0 + 0.5 * (gamma_a - 1.0) * Mach2 ** 2)
            p2 = (IPR * p0_isa) * (T2 / Tt2) ** (gamma_a / (gamma_a - 1.0))
            rho2 = p2 / (r_gas * T2)
            V2 = Mach2 * math.sqrt(gamma_a * r_gas * T2)
            A_required = mass_flow / (rho2 * V2)

            if A_required > A_geom:
                errors.append(
                    f"ERROR: Required flow area A_req={A_required:.4f} m² > "
                    f"geometric area A_geom={A_geom:.4f} m² at d_max={d_max:.3f} m."
                )
        except Exception as e:
            errors.append(
                f"WARNING: Could not run mass-flow vs d_max check: {e}"
            )

        return errors

    def _write_defaults(self, path):
        """Helper to write default values to the specified path."""
        self._save_data(
            path,
            _DEFAULT_FLIGHT,
            _DEFAULT_FEATURES,
            _DEFAULT_GEOMETRY,
            _DEFAULT_MATERIALS,
        )

    def launch_gui(self, filepath=None):
        import shutil
        import os

        if filepath:
            self.filepath = filepath

        # Resolve session file path
        if self.work_dir:
            session_path = os.path.join(self.work_dir, SESSION_FILENAME)

            # Copy template to work_dir if session file doesn't exist yet
            if not os.path.exists(session_path):
                if self.filepath and os.path.exists(self.filepath):
                    shutil.copy2(self.filepath, session_path)
                else:
                    # No template: create default xlsx from schema
                    self._write_defaults(session_path)
            self._session_path = session_path
        else:
            self._session_path = None

        self._pending_l2_errors = []

        while True:
            result = self._run_tkinter_ui()   # always reads/writes self._session_path
            if result == "cancelled":
                return None

            # Level 2 validation — read directly from disk, no cache
            raw = self._load(self._session_path)
            loaded_flight, loaded_features, loaded_geometry, loaded_materials = raw

            flight = {**_DEFAULT_FLIGHT, **loaded_flight}
            features = {**_DEFAULT_FEATURES, **loaded_features}
            geometry = {**_DEFAULT_GEOMETRY, **loaded_geometry}
            materials = {**_DEFAULT_MATERIALS, **loaded_materials}

            errors   = self.validate(flight, features, geometry, materials)
            warnings = [e for e in errors if e.startswith("WARNING")]
            level2   = [e for e in errors if e.startswith("ERROR")]

            for w in warnings:
                print(f"[InputParser] {w}")

            if not level2:
                # Validation passed: point filepath at session file
                # This is a REAL value change -> ParaPy cache busted cleanly
                if self._session_path:
                    self.filepath = self._session_path
                break

            print("[InputParser] Cross-parameter validation failed:")
            for e in level2:
                print(f"  {e}")
            print("[InputParser] Re-opening UI for correction...")
            self._pending_l2_errors = level2

        return (self.filepath, self.work_dir, flight, features, geometry, materials)

    def _run_tkinter_ui(self):
        import tkinter as tk
        from tkinter import ttk, filedialog, messagebox
        import os

        was_filepath_set = bool(self.filepath)

        # Load existing data from the session file
        loaded_flight, loaded_features, loaded_geometry, loaded_materials = ({}, {}, {}, {})
        if hasattr(self, "_session_path") and self._session_path and os.path.exists(self._session_path):
            loaded_flight, loaded_features, loaded_geometry, loaded_materials = self._load(self._session_path)

        root = tk.Tk()
        root.title("AeroEngine Design Configurator")
        root.geometry("900x750")
        root.configure(bg="#20232a")

        # Dark theme Styling
        style = ttk.Style(root)
        style.theme_use('clam')

        style.configure(".", background="#20232a", foreground="#ffffff", font=("Segoe UI", 10))
        style.configure("TNotebook", background="#20232a", borderwidth=0)
        style.configure("TNotebook.Tab", background="#282c34", foreground="#abb2bf", font=("Segoe UI", 10, "bold"), padding=[12, 6])
        style.map("TNotebook.Tab", background=[("selected", "#20232a")], foreground=[("selected", "#61dafb")])
        style.configure("TFrame", background="#20232a")
        style.configure("TLabel", background="#20232a", foreground="#ffffff")
        style.configure("TCombobox", fieldbackground="#353b45", foreground="#ffffff", selectbackground="#353b45")

        style.configure("Primary.TButton", background="#4a90e2", foreground="#ffffff", font=("Segoe UI", 10, "bold"), borderwidth=0, padding=8)
        style.map("Primary.TButton", background=[("active", "#357abd")])
        style.configure("Secondary.TButton", background="#4f5b66", foreground="#ffffff", font=("Segoe UI", 10), borderwidth=0, padding=8)
        style.map("Secondary.TButton", background=[("active", "#65737e")])

        # Header Title
        title_label = ttk.Label(root, text="AeroEngine KBE Input Parser", font=("Segoe UI", 16, "bold"), foreground="#61dafb")
        title_label.pack(pady=15)

        # Wizard container to toggle between Folder page and Parameter tabs
        wizard_container = ttk.Frame(root)
        wizard_container.pack(fill="both", expand=True, padx=15, pady=5)

        notebook = ttk.Notebook(wizard_container)

        # We will store variables for each category
        vars_flight = {}
        vars_features = {}
        vars_geometry = {}
        vars_materials = {}

        material_choices = list(MATERIAL_DB.keys())

        # Tab 0: Flight Conditions
        tab_flight = ttk.Frame(notebook)
        notebook.add(tab_flight, text="Flight Conditions")

        # Tab 1: Cycle & Performance
        tab_features = ttk.Frame(notebook)
        notebook.add(tab_features, text="Cycle & Performance")

        # Tab 2: Geometry
        tab_geometry = ttk.Frame(notebook)
        notebook.add(tab_geometry, text="Geometry")

        # Tab 3: Materials
        tab_materials = ttk.Frame(notebook)
        notebook.add(tab_materials, text="Materials")

        # Validation logic state
        self._field_valid = {}
        entry_widgets = {}

        def _validate_field_live(entry_widget, key, section, value_str):
            schema_entry = INPUT_SCHEMA[section][key]
            try:
                val = float(value_str)
                if schema_entry["min"] <= val <= schema_entry["max"]:
                    entry_widget.config(bg="#353b45")   # normal background
                    return True
                else:
                    entry_widget.config(bg="#5a0000")   # dark red = out of range
                    return False
            except ValueError:
                entry_widget.config(bg="#5a0000")
                return False

        def _update_save_button_state():
            all_valid = all(self._field_valid.values())
            save_btn.config(state="normal" if all_valid else "disabled")

        def make_validate_cmd(e_widget, k, sect, v_var):
            def handler(event):
                val_str = v_var.get().strip()
                is_valid = _validate_field_live(e_widget, k, sect, val_str)
                self._field_valid[(sect, k)] = is_valid
                _update_save_button_state()
            return handler

        # Populate Tab 0: Flight Conditions (Single column)
        flight_frame = ttk.Frame(tab_flight)
        flight_frame.pack(padx=20, pady=20, fill="both", expand=True)
        for i, (key, item) in enumerate(INPUT_SCHEMA["design_flight_conditions"].items()):
            val = loaded_flight.get(key, item["default"])
            var = tk.StringVar(value=str(val))
            vars_flight[key] = var

            lbl = ttk.Label(flight_frame, text=item["label"] + ":")
            lbl.grid(row=i, column=0, sticky="w", padx=10, pady=8)

            ent = tk.Entry(
                flight_frame,
                textvariable=var,
                width=15,
                font=("Segoe UI", 10),
                bg="#353b45",
                fg="#ffffff",
                insertbackground="white",
                relief="flat",
                bd=1
            )
            ent.grid(row=i, column=1, padx=10, pady=8, sticky="w")
            entry_widgets[("design_flight_conditions", key)] = ent

            unit_text = f"{item['unit']}  [{item['min']} to {item['max']}]"
            unit_lbl = ttk.Label(flight_frame, text=unit_text, foreground="#abb2bf")
            unit_lbl.grid(row=i, column=2, padx=10, pady=8, sticky="w")

            # Initial bounds check validation
            self._field_valid[("design_flight_conditions", key)] = _validate_field_live(
                ent, key, "design_flight_conditions", str(val)
            )
            ent.bind("<KeyRelease>", make_validate_cmd(ent, key, "design_flight_conditions", var))

        # Populate Tab 1: Cycle & Performance (Two columns)
        features_frame = ttk.Frame(tab_features)
        features_frame.pack(padx=20, pady=20, fill="both", expand=True)
        feat_keys = list(INPUT_SCHEMA["engine_features"].keys())
        half_feat = (len(feat_keys) + 1) // 2

        for idx, key in enumerate(feat_keys):
            item = INPUT_SCHEMA["engine_features"][key]
            val = loaded_features.get(key, item["default"])
            var = tk.StringVar(value=str(val))
            vars_features[key] = var

            col_offset = 0 if idx < half_feat else 4
            row_idx = idx if idx < half_feat else idx - half_feat

            lbl = ttk.Label(features_frame, text=item["label"] + ":")
            lbl.grid(row=row_idx, column=col_offset, sticky="w", padx=5, pady=4)

            ent = tk.Entry(
                features_frame,
                textvariable=var,
                width=10,
                font=("Segoe UI", 10),
                bg="#353b45",
                fg="#ffffff",
                insertbackground="white",
                relief="flat",
                bd=1
            )
            ent.grid(row=row_idx, column=col_offset+1, padx=5, pady=4, sticky="w")
            entry_widgets[("engine_features", key)] = ent

            unit_text = f"{item['unit']} [{item['min']}-{item['max']}]"
            unit_lbl = ttk.Label(features_frame, text=unit_text, foreground="#abb2bf", font=("Segoe UI", 9))
            unit_lbl.grid(row=row_idx, column=col_offset+2, padx=5, pady=4, sticky="w")

            # Spacer column between the two halves
            if col_offset == 0:
                spacer = ttk.Label(features_frame, text="   ")
                spacer.grid(row=row_idx, column=3)

            # Initial bounds check validation
            self._field_valid[("engine_features", key)] = _validate_field_live(
                ent, key, "engine_features", str(val)
            )
            ent.bind("<KeyRelease>", make_validate_cmd(ent, key, "engine_features", var))

        # Populate Tab 2: Geometry (Two columns)
        geom_frame = ttk.Frame(tab_geometry)
        geom_frame.pack(padx=20, pady=20, fill="both", expand=True)
        geom_keys = list(INPUT_SCHEMA["engine_geometry"].keys())
        half_geom = (len(geom_keys) + 1) // 2

        for idx, key in enumerate(geom_keys):
            item = INPUT_SCHEMA["engine_geometry"][key]
            val = loaded_geometry.get(key, item["default"])
            var = tk.StringVar(value=str(val))
            vars_geometry[key] = var

            col_offset = 0 if idx < half_geom else 4
            row_idx = idx if idx < half_geom else idx - half_geom

            lbl = ttk.Label(geom_frame, text=item["label"] + ":")
            lbl.grid(row=row_idx, column=col_offset, sticky="w", padx=8, pady=6)

            ent = tk.Entry(
                geom_frame,
                textvariable=var,
                width=12,
                font=("Segoe UI", 10),
                bg="#353b45",
                fg="#ffffff",
                insertbackground="white",
                relief="flat",
                bd=1
            )
            ent.grid(row=row_idx, column=col_offset+1, padx=8, pady=6, sticky="w")
            entry_widgets[("engine_geometry", key)] = ent

            unit_text = f"{item['unit']} [{item['min']}-{item['max']}]"
            unit_lbl = ttk.Label(geom_frame, text=unit_text, foreground="#abb2bf")
            unit_lbl.grid(row=row_idx, column=col_offset+2, padx=8, pady=6, sticky="w")

            if col_offset == 0:
                spacer = ttk.Label(geom_frame, text="     ")
                spacer.grid(row=row_idx, column=3)

            # Initial bounds check validation
            self._field_valid[("engine_geometry", key)] = _validate_field_live(
                ent, key, "engine_geometry", str(val)
            )
            ent.bind("<KeyRelease>", make_validate_cmd(ent, key, "engine_geometry", var))

        # Populate Tab 3: Materials (Single column with Comboboxes)
        materials_frame = ttk.Frame(tab_materials)
        materials_frame.pack(padx=20, pady=20, fill="both", expand=True)
        for i, (key, item) in enumerate(INPUT_SCHEMA["engine_materials"].items()):
            val = loaded_materials.get(key, item["default"])
            var = tk.StringVar(value=str(val))
            vars_materials[key] = var

            lbl = ttk.Label(materials_frame, text=item["label"] + ":")
            lbl.grid(row=i, column=0, sticky="w", padx=10, pady=10)

            cb = ttk.Combobox(materials_frame, textvariable=var, values=material_choices, width=25, state="readonly")
            cb.grid(row=i, column=1, padx=10, pady=10, sticky="w")

            unit_lbl = ttk.Label(materials_frame, text=item["unit"], foreground="#abb2bf")
            unit_lbl.grid(row=i, column=2, padx=10, pady=10, sticky="w")
            
            self._field_valid[("engine_materials", key)] = True

        # Error text area at the bottom
        bottom_frame = ttk.Frame(root)

        error_title = ttk.Label(bottom_frame, text="Validation Errors:", font=("Segoe UI", 10, "bold"), foreground="#ff5555")
        error_title.pack(anchor="w")

        error_text = tk.Text(bottom_frame, height=5, fg="#ff5555", bg="#1e1e24", insertbackground="white", font=("Courier New", 9))
        error_text.pack(fill="x", pady=5)
        error_text.config(state="disabled")

        # Action Buttons frame
        buttons_frame = ttk.Frame(bottom_frame)
        buttons_frame.pack(anchor="e", pady=5)

        gui_result = ["cancelled"]  # default is cancelled

        # ------------------ Wizard Page 1: Folder Selection ------------------
        folder_frame = ttk.Frame(wizard_container)
        folder_frame.pack(fill="both", expand=True, padx=20, pady=20)

        # Instructions / Section Header
        lbl_info = ttk.Label(folder_frame, text="Step 1 — Project Configuration Paths", font=("Segoe UI", 12, "bold"), foreground="#61dafb")
        lbl_info.grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 20))

        # Row 1: Excel Template File
        template_lbl = ttk.Label(folder_frame, text="Template Excel Path:")
        template_lbl.grid(row=1, column=0, sticky="w", padx=5, pady=10)

        template_var = tk.StringVar(value=self.filepath or "")
        template_ent = ttk.Entry(folder_frame, textvariable=template_var, width=65, state="readonly")
        template_ent.grid(row=1, column=1, padx=10, pady=10, sticky="ew")

        if was_filepath_set:
            lock_lbl = ttk.Label(folder_frame, text="🔒 (locked)", font=("Segoe UI", 10, "italic"), foreground="#ff5555")
            lock_lbl.grid(row=1, column=2, padx=5, pady=10, sticky="w")
        else:
            def browse_template():
                selected = filedialog.askopenfilename(
                    parent=root,
                    title="Select Excel Template File",
                    filetypes=[("Excel Files", "*.xlsx")]
                )
                if selected:
                    template_var.set(selected)
                    self.filepath = selected

            browse_temp_btn = ttk.Button(folder_frame, text="Browse...", command=browse_template, style="Secondary.TButton")
            browse_temp_btn.grid(row=1, column=2, padx=5, pady=10, sticky="w")

        # Row 2: Working Directory
        work_dir_lbl = ttk.Label(folder_frame, text="Working Directory:")
        work_dir_lbl.grid(row=2, column=0, sticky="w", padx=5, pady=10)

        work_dir_var = tk.StringVar(value=self.work_dir)
        work_dir_ent = ttk.Entry(folder_frame, textvariable=work_dir_var, width=65, state="readonly")
        work_dir_ent.grid(row=2, column=1, padx=10, pady=10, sticky="ew")

        def browse_work_dir():
            selected = filedialog.askdirectory(
                parent=root,
                title="Select Working Directory Folder"
            )
            if selected:
                work_dir_var.set(selected)
                self.work_dir = selected

        browse_wd_btn = ttk.Button(folder_frame, text="Browse...", command=browse_work_dir, style="Secondary.TButton")
        browse_wd_btn.grid(row=2, column=2, padx=5, pady=10, sticky="w")

        folder_frame.columnconfigure(1, weight=1)

        # Next button
        nav_buttons_frame = ttk.Frame(folder_frame)
        nav_buttons_frame.grid(row=3, column=1, columnspan=2, sticky="e", pady=30)

        def on_next():
            if not self.work_dir:
                messagebox.showwarning("Warning", "Please select a working directory folder before proceeding.", parent=root)
                return

            # Hide Folder frame
            folder_frame.pack_forget()

            # Resolve the session path
            session_path = os.path.join(self.work_dir, SESSION_FILENAME)
            self._session_path = session_path

            # Copy selected template to session file if it doesn't exist yet
            if not os.path.exists(session_path):
                import shutil
                if self.filepath and os.path.exists(self.filepath):
                    shutil.copy2(self.filepath, session_path)
                else:
                    self._write_defaults(session_path)

            # Load selected template data if set
            if self._session_path and os.path.exists(self._session_path):
                loaded_flight_next, loaded_features_next, loaded_geometry_next, loaded_materials_next = self._load(self._session_path)

                # Populate / update tab fields with loaded data
                for k, var in vars_flight.items():
                    val = loaded_flight_next.get(k, INPUT_SCHEMA["design_flight_conditions"][k]["default"])
                    var.set(str(val))
                    ent = entry_widgets[("design_flight_conditions", k)]
                    self._field_valid[("design_flight_conditions", k)] = _validate_field_live(
                        ent, k, "design_flight_conditions", str(val)
                    )

                for k, var in vars_features.items():
                    val = loaded_features_next.get(k, INPUT_SCHEMA["engine_features"][k]["default"])
                    var.set(str(val))
                    ent = entry_widgets[("engine_features", k)]
                    self._field_valid[("engine_features", k)] = _validate_field_live(
                        ent, k, "engine_features", str(val)
                    )

                for k, var in vars_geometry.items():
                    val = loaded_geometry_next.get(k, INPUT_SCHEMA["engine_geometry"][k]["default"])
                    var.set(str(val))
                    ent = entry_widgets[("engine_geometry", k)]
                    self._field_valid[("engine_geometry", k)] = _validate_field_live(
                        ent, k, "engine_geometry", str(val)
                    )

                for k, var in vars_materials.items():
                    var.set(str(loaded_materials_next.get(k, INPUT_SCHEMA["engine_materials"][k]["default"])))

                _update_save_button_state()

            # Show notebooks & buttons
            notebook.pack(fill="both", expand=True, padx=15, pady=5)
            bottom_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=15, pady=15)

        next_btn = ttk.Button(nav_buttons_frame, text="Next →", command=on_next, style="Primary.TButton")
        next_btn.pack(side="right", padx=5)

        # ------------------ Existing Tab Callback Actions ------------------
        def on_save():
            # Clear previous errors
            error_text.config(state="normal")
            error_text.delete("1.0", tk.END)
            error_text.config(state="disabled")

            flight = {}
            features = {}
            geometry = {}
            materials = {}

            # Read variables and convert types
            # Flight Conditions
            for k, var in vars_flight.items():
                flight[k] = float(var.get().strip())

            # Engine Features
            for k, var in vars_features.items():
                features[k] = float(var.get().strip())

            # Engine Geometry
            for k, var in vars_geometry.items():
                geometry[k] = float(var.get().strip())

            # Engine Materials
            for k, var in vars_materials.items():
                materials[k] = var.get().strip()

            # Save the newly entered values to the session Excel file
            self._save_data(self._session_path, flight, features, geometry, materials)

            # Store the result to return
            gui_result[0] = "saved"
            root.destroy()

        def on_cancel():
            root.destroy()

        save_btn = ttk.Button(buttons_frame, text="Save & Close", command=on_save, style="Primary.TButton")
        save_btn.pack(side="left", padx=5)

        # Initial Save button state (depends on loaded values validity)
        _update_save_button_state()

        cancel_btn = ttk.Button(buttons_frame, text="Cancel", command=on_cancel, style="Secondary.TButton")
        cancel_btn.pack(side="left", padx=5)

        root.protocol("WM_DELETE_WINDOW", on_cancel)

        # On startup, check for L2 errors
        if getattr(self, "_pending_l2_errors", None):
            folder_frame.pack_forget()
            notebook.pack(fill="both", expand=True, padx=15, pady=5)
            bottom_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=15, pady=15)

            # Show L2 errors
            error_text.config(state="normal")
            error_text.delete("1.0", tk.END)
            for err in self._pending_l2_errors:
                error_text.insert(tk.END, f"• {err}\n")
            error_text.config(state="disabled")
            self._pending_l2_errors = []

        root.mainloop()

        return gui_result[0]


if __name__ == "__main__":
    parser = InputParser()
    print("Testing launch_gui with defaults...")
    # This will open the GUI, prefilled with defaults since filepath is empty.
    # When user clicks Save & Close, it will ask where to save.
    result = parser.launch_gui()
    print(f"GUI returned: {result}")

    if result:
        filepath = result[0]
        # Instantiate a new parser with the written path and verify values
        print(f"Loading written file: {filepath}")
        new_parser = InputParser(filepath=filepath)
        print(f"Mach condition: {new_parser.flight_conditions['Mach']} (Expected from GUI)")

        # Test validation on invalid inputs
        print("Testing validation on invalid inputs...")
        errors1 = new_parser.validate(
            flight={**new_parser.flight_conditions, "Mach": 1.5},  # out of range [0.0, 1.0]
            features={**new_parser.engine_features, "C_eta": 1.5},  # out of range [0.80, 1.0]
            geometry={**new_parser.engine_geometry, "spool_tip_length": 2.0, "spool_length": 1.5},  # tip >= length
            materials={**new_parser.engine_materials, "casing": "Unobtainium"}  # not in DB
        )
        print("Validation errors:")
        for err in errors1:
            print(f"  {err}")
